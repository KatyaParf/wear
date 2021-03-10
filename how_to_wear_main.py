#
# базовый пример с детектированием объектов (люди, кошки)
#

from datetime import datetime
import pathlib
import tensorflow as tf
import numpy as np
from PIL import Image
from object_detection.utils import ops as utils_ops
from object_detection.utils import label_map_util
from object_detection.utils import visualization_utils as vis_util

import sys
import operator
import array
import openpyxl

#----------

class Wear(object):
    def __init__(self, aName, aClass):
        self.name = aName
        self.cls = aClass

class Pogoda(object):
    def __init__(self, aMinTemp, aMaxTemp, aOsad):
        self.min_temp = aMinTemp
        self.max_temp = aMaxTemp
        self.osad = aOsad

# для хранения списка одежды (по группам) для интервала погоды
class IntervalInfo(object):
    def __init__(self):
        self.pogoda = None
        # словарь, ключ = группа, значение = массив названий одежды
        self.wear_names = {}

# для хранения имени и группы одежды и списка погод для класса
class ClassInfo(object):
    def __init__(self):
        self.wear_name = ""
        self.grp = ""
        self.pogodas = []

#----------

# массив интервалов температур с подходящей одеждой
intervals = []

# разные классы одежды ????
MapByClass = {}

#----------

# считывание excel файла
def ReadConfig():
    wb = openpyxl.load_workbook("what_to_wear.xlsx")
    sheet = wb.active

    # read count of intervals
    cols = sheet.cell(row = 1, column = 1).value

    tempwears = []

    aRow = 5
    # считываем классы
    while True:
        wear_name  = sheet.cell(row = aRow, column = 1).value
        wear_class = sheet.cell(row = aRow, column = 2).value
        if (wear_name == "*end"):
            break
        tempwears.append(Wear(wear_name, wear_class))
        aRow = aRow + 1

    # считываем интервалы температур
    for col in range(1 + 2, cols + 2 + 1):
        newInterval = IntervalInfo()
        newInterval.pogoda = Pogoda(
            sheet.cell(row = 2, column = col).value,
            sheet.cell(row = 3, column = col).value,
            sheet.cell(row = 4, column = col).value )
        # добавляем интервал в общий список
        intervals.append(newInterval)

        lastgrp = ''
        # идем по строкам таблицы
        for aRow in range(0, len(tempwears)):
            wear_name = tempwears[aRow].name

            # встретили строку - группу, запомнили ее
            if (wear_name[0] == "*"):
                lastgrp = wear_name

            # строка - не группа
            cell = sheet.cell(row = aRow + 5, column = col).value
            # стоит отметка что одежда подходит для температуры
            if (cell == 1):
                # добавляем группу (ключ) в словарь одежды для интервала
                if not (lastgrp in newInterval.wear_names):
                    newInterval.wear_names[lastgrp] = []
                # в эту группу добавляем одежду
                newInterval.wear_names[lastgrp].append(wear_name)


                # добавляем класс одежды в глобальный список классов
                wear_class = tempwears[aRow].cls
                if not (wear_class in MapByClass):
                    newClass = ClassInfo()
                    newClass.wear_name = wear_name
                    newClass.grp = lastgrp
                    MapByClass[wear_class] = newClass

                pogoda = Pogoda(
                    newInterval.pogoda.min_temp,
                    newInterval.pogoda.max_temp,
                    newInterval.pogoda.osad)

                MapByClass[wear_class].pogodas.append(pogoda)

#------------------
#на вход - массив с англ классами одежды, температура и осадки в наст время
#на выход - массив с не подходящими классами одежды
def get_wrong_wear(aClasses, temp, osad):
    res = []

    for className in aClasses:
        classInfo = MapByClass[className]

        found = False
        for p in classInfo.pogodas:
            if (p.min_temp <= temp) and (p.max_temp > temp) and ((osad == p.osad) or (p.osad == 2)):
                found = True
                break
        if found == False:
            res.append(className)

    return res


# анализ того что надето неверно и совет что надеть
# wrong_classes - список неверных классов предметов одежды (массив строк)
def sovet(wrong_classes, temp, osad):
    if len(wrong_classes) == 0:
        return "Вы одеты идеально!"

    res_str = "На вас надето: "
    # предполагаем что одновременно надето не более одной одежды из каждого
    # класса
    # перебираем все полученнные классы одежды,
    for className in wrong_classes:
        classInfo = MapByClass[className]
        wearName = classInfo.wear_name
        res_str += wearName + ", "
    
    res_str = res_str[:-2] + ". Вместо этого я советую вам надеть: "

    for className in wrong_classes:
        classInfo = MapByClass[className]

        # получаем группу для класса
        grp = classInfo.grp

        # для полученной группы, надо выводить список замен одежды для нужной
        # температуры
        # в формате: имя-группы: список предметов из этой группы для нужной
        # температуры
        # например: На голову: Легкая вязаная шапка или легкий платок
        for i in intervals:
            if (i.pogoda.min_temp <= temp) and (i.pogoda.max_temp > temp) and ((osad == i.pogoda.osad) or (i.pogoda.osad == 2)):
                # список одежды из группы
                for w in i.wear_names[grp]:
                    res_str += w + " или "
                res_str = res_str[:-5] + ", "

    res_str = res_str[:-2] + "."
    return res_str

#------------------



def load_model(model_name):
  model_dir = model_name + "/saved_model"
  model = tf.saved_model.load(model_dir)
  return model

# Загрузка модели для выделения тела
model_name = '_my_models/body_model' # самая маленькая и быстрая. качество - нормально находит людей
detection_model_body = load_model(model_name)

# Загрузка модели для выделения одежды
model_name = '_my_models/wear_model' # находит одежду
detection_model_wear = load_model(model_name)

#------------------

# Список с названиями для каждого класса в модели
PATH_TO_LABELS = '_my_models/wear_labels/label_map.pbtxt'
category_index_wear = label_map_util.create_category_index_from_labelmap(PATH_TO_LABELS, use_display_name=True)

#----
# Основная функция обработки фотографии
def process_image_body(model_body, model_wear, image_path, dst_folder, postfix):
  image_src = Image.open(image_path)

  #---------------------------------
  #--- выделение тела на фото ------

  image_np = np.array(image_src)
  # The input needs to be a tensor, convert it using `tf.convert_to_tensor`.
  image_arr = np.asarray(image_np)
  input_tensor = tf.convert_to_tensor(image_arr)
  # The model expects a batch of images, so add an axis with `tf.newaxis`.
  input_tensor2 = input_tensor[tf.newaxis,...]

  # Run inference
  model_fn = model_body.signatures['serving_default']
  output_dict = model_fn(input_tensor2)

  num_detections = int(output_dict.pop('num_detections'))
  output_dict = {key:value[0, :num_detections].numpy()
                 for key,value in output_dict.items()}
  output_dict['num_detections'] = num_detections
  output_dict['detection_classes'] = output_dict['detection_classes'].astype(np.int64)

  #----

  # записываем обрезанное изображение с фигурой
  height, width = image_np.shape[0], image_np.shape[1]
  for i, box in enumerate(output_dict['detection_boxes']):
    # get class: 1 with score > 0.6
    if (output_dict['detection_classes'][i] == 1) and (output_dict['detection_scores'][i] > 0.6):
        (xmin, ymin, xmax, ymax) = (int(box[1]*width), int(box[0]*height), int(box[3]*width), int(box[2]*height))
        im_crop = image_src.crop((xmin, ymin, xmax, ymax));


  #---------------------------------
  #--- выделение одежды на фото ------

  image_np = np.array(im_crop)

  # The input needs to be a tensor, convert it using `tf.convert_to_tensor`.
  image_arr = np.asarray(image_np)
  input_tensor = tf.convert_to_tensor(image_arr)
  # The model expects a batch of images, so add an axis with `tf.newaxis`.
  input_tensor2 = input_tensor[tf.newaxis,...]

  # Run inference
  model_fn = model_wear.signatures['serving_default']
  output_dict = model_fn(input_tensor2)

  num_detections = int(output_dict.pop('num_detections'))
  output_dict = {key:value[0, :num_detections].numpy()
                 for key,value in output_dict.items()}
  output_dict['num_detections'] = num_detections
  output_dict['detection_classes'] = output_dict['detection_classes'].astype(np.int64)

  #----
  # remove found ignored classes
  classes = []
  ignored_classes = [69, 228, 502, 221, 308]
  for i, cls in enumerate(output_dict['detection_classes']):
    if cls in ignored_classes:
        output_dict['detection_scores'][i] = 0.01
    else:
        #print ("   ",
        #    category_index_wear[output_dict['detection_classes'][i]]['name'],
        #    "{:.2f}".format(output_dict['detection_scores'][i])
        #)
        classes.append(category_index_wear[output_dict['detection_classes'][i]]['name'])

  return classes

#------------------

ReadConfig()

classes = process_image_body(detection_model_body, detection_model_wear,
  pathlib.Path("_door_photo.jpg"), ".", "_wear")
print(classes)

sredTemp = int(float(sys.argv[1]))
sredOsad = int(sys.argv[2])

res = get_wrong_wear(classes, sredTemp, sredOsad)
print(res)
sov = sovet(res, sredTemp, sredOsad)
print(sov)
