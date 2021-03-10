import sys
import operator
import array

import openpyxl
#from openpyxl.reader.excel import load_workbook

#----------

class Wear(object):
    def __init__(self, aName, aClass):
        self.name = aName
        self.cls = aClass

class Pogoda(object):
    def __init__(self, aMinTemp, aMaxTemp, aOsad):
        self.min_temp = aMinTemp
        self.max_temp = aMaxTemp
        self.osad = aOsad

# для хранения списка одежды (по группам) для интервала погоды
class IntervalInfo(object):
    def __init__(self):
        self.pogoda = None
        # словарь, ключ = группа, значение = массив названий одежды
        self.wear_names = {}

#----------

# массив интервалов температур с подходящей одеждой
intervals = []

#----------

# считывание excel файла
def ReadConfig():
    wb = openpyxl.load_workbook("what_to_wear.xlsx")
    sheet = wb.active

    # read count of intervals
    cols = sheet.cell(row = 1, column = 1).value

    tempwears = []

    aRow = 5
    # считываем классы
    while True:
        wear_name  = sheet.cell(row = aRow, column = 1).value
        wear_class = sheet.cell(row = aRow, column = 2).value
        if (wear_name == "*end"):
            break
        tempwears.append(Wear(wear_name, wear_class))
        aRow = aRow + 1

    # считываем интервалы температур
    for col in range(1 + 2, cols + 2 + 1):
        newInterval = IntervalInfo()
        newInterval.pogoda = Pogoda(
            sheet.cell(row = 2, column = col).value,
            sheet.cell(row = 3, column = col).value,
            sheet.cell(row = 4, column = col).value )
        # добавляем интервал в общий список
        intervals.append(newInterval)

        lastgrp = ''
        # идем по строкам таблицы
        for aRow in range(0, len(tempwears)):
            wear_name = tempwears[aRow].name

            # встретили строку - группу, запомнили ее
            if (wear_name[0] == "*"):
                lastgrp = wear_name

            # строка - не группа
            cell = sheet.cell(row = aRow + 5, column = col).value
            # стоит отметка что одежда подходит для температуры
            if (cell == 1):
                # добавляем группу (ключ) в словарь одежды для интервала
                if not (lastgrp in newInterval.wear_names):
                    newInterval.wear_names[lastgrp] = []
                # в эту группу добавляем одежду
                newInterval.wear_names[lastgrp].append(wear_name)


#-----------

# выдача рекомендаций по заданной темп и осадкам
def recommend(temp, osad):

    res_str = "Сегодня температура: " + str(temp)
    if osad == 1 and temp <= 0:
        res_str += ", ожидается снег"
    elif osad == 1 and temp > 0:
        res_str += ", ожидается дождь"
    res_str += ". "
    res_str += "Советую надеть: "

    # находим нужные столбцы
    # формируем набор предметов одежды по всем столбцам
    for i in intervals:
        if (i.pogoda.min_temp <= temp) and (i.pogoda.max_temp > temp) and ((osad == i.pogoda.osad) or (i.pogoda.osad == 2)):
            # проходим по всем группам интервала
            for grp in i.wear_names.keys():
                # пишем имя группы
                res_str += "\n   " + grp[1:] + ": "
                # и список одежды из группы
                for w in i.wear_names[grp]:
                    res_str += w + " или "
                res_str = res_str[:-5] + ". "

    return res_str


#------------------

def main(argv):
    ReadConfig()

    sredTemp = int(float(sys.argv[1]))
    sredOsad = int(sys.argv[2])

    res = recommend(sredTemp, sredOsad)
    print(res)

#-----------

if __name__ == "__main__":
    main(sys.argv)